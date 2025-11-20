import streamlit as st
import pandas as pd
import os
import io
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns

# ---------- CONFIG ----------
ACTIVE_SHEET = "Active"
RESIGNED_SHEET = "Resigned-Contract End"
EXCEL_FILE = "staff_data.xlsx"
PROFILE_IMG_DIR = "profile_images"
PROJECTS = ["Afghan response", "Flood response", "Flow Monitoring", "PMS", "EMS"]

os.makedirs(PROFILE_IMG_DIR, exist_ok=True)
# ---------- LOGIN ----------
USERNAME = st.secrets["login"]["username"]
PASSWORD = st.secrets["login"]["password"]

def login():
    st.title("🔐 Login")
    if "password_verified" not in st.session_state:
        st.session_state.password_verified = False

    username_input = st.text_input("Username")
    password_input = st.text_input("Password", type="password")

    if st.button("Unlock"):
        # Check against secrets
        if username_input == st.secrets["login"]["username"] and password_input == st.secrets["login"]["password"]:
            st.session_state.password_verified = True
            st.success("✅ Login successful. Loading app...")
            st.rerun()
        else:
            st.error("❌ Incorrect username or password.")


# ---------- UTILS ----------
def load_data():
    try:
        xls = pd.ExcelFile(EXCEL_FILE)
        active_df = pd.read_excel(xls, ACTIVE_SHEET)
        resigned_df = pd.read_excel(xls, RESIGNED_SHEET)

        col_map = {
            "Designation_Name": "Designation",
            "Department/Unit": "Unit",
            "Place_of_Posting_Location": "District - Duty Station",
            "Starting_Salary": "Starting_Salary (PKR)"
        }
        resigned_df.rename(columns=col_map, inplace=True)

        for df in [active_df, resigned_df]:
            for project in PROJECTS:
                if project not in df.columns:
                    df[project] = False
            if "Profile_Image" not in df.columns:
                df["Profile_Image"] = ""

        return active_df, resigned_df

    except Exception as e:
        st.error(f"❌ Failed to load staff data: {e}")
        return pd.DataFrame(), pd.DataFrame()

def save_data(active_df, resigned_df):
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
        active_df.to_excel(writer, sheet_name=ACTIVE_SHEET, index=False)
        resigned_df.to_excel(writer, sheet_name=RESIGNED_SHEET, index=False)

# ---------- APP ----------
if "password_verified" not in st.session_state or not st.session_state.password_verified:
    login()
    st.stop()

# ---------- PAGE CONFIG ----------
st.set_page_config(page_title="HR Dashboard", layout="wide")

# Add a logout button to the sidebar
with st.sidebar:
    if st.button("🔓 Logout"):
        st.session_state.clear()
        st.rerun()

# ---------- REST OF YOUR APP ----------
# (Paste your complete app code from "# ---------- MENU ----------" onwards here)

# Example of loading data and showing the dashboard tab
active_df, resigned_df = load_data()



# ---------- PAGE CONFIG ----------
st.set_page_config(page_title="HR Dashboard", layout="wide")

# ---------- MENU ----------
# Initialize selected tab in session state
if "menu" not in st.session_state:
    st.session_state.menu = "🏠 Dashboard"

# List of tabs
tabs = [
    "🏠 Dashboard",
    "👥 View Profiles",
    "✏️ Edit Employee",
    "📤 Close Contract",
    "➕ Add Staff",
    "❌ Delete Staff",
    "📥 Download Data",
    "🚫 Inactive Staff",
     "📆 Attendance"
]

# Sidebar UI (replaces st.sidebar.radio)
with st.sidebar:
    for tab in tabs:
        is_selected = (st.session_state.menu == tab)
        style = """
            background-color: #4A90E2;
            color: white;
            font-weight: bold;
        """ if is_selected else """
            background-color: transparent;
            color: black;
        """
        button = st.button(tab, key=tab, use_container_width=True)
        st.markdown(
            f"""
            <style>
                div[data-testid="stButton"][key="{tab}"] > button {{
                    border: none;
                    text-align: left;
                    padding: 0.75rem 1rem;
                    margin-bottom: 0.5rem;
                    border-radius: 8px;
                    {style}
                }}
            </style>
            """,
            unsafe_allow_html=True
        )
        if button:
            st.session_state.menu = tab
            st.rerun()

# Use selected tab
menu = st.session_state.menu

active_df, resigned_df = load_data()

if menu == "🏠 Dashboard":
    st.title("📊 HR Dashboard Overview")

    # ───── KPI Cards ─────
    total_active = len(active_df)
    total_resigned = len(resigned_df)
    today = pd.to_datetime(datetime.today())
    active_df['Contract_End_Date'] = pd.to_datetime(active_df['Contract_End_Date'], errors='coerce')
    contracts_expiring = active_df[(active_df['Contract_End_Date'] - today).dt.days.between(0, 30)].shape[0]

    k1, k2, k3 = st.columns(3)
    k1.metric("👥 Active Staff", total_active)
    k2.metric("📤 Resigned Staff", total_resigned)
    k3.metric("⚠️ Expiring in 30 Days", contracts_expiring)

    st.markdown("---")
     # ───── Contracts Expiring Soon ─────
    st.subheader("📅 Contract Expiry Alerts")

    upcoming = active_df.copy()
    upcoming = upcoming.dropna(subset=['Contract_End_Date'])
    upcoming['Contract_End_Date'] = pd.to_datetime(upcoming['Contract_End_Date'], errors='coerce')

    # Today's date
    today = pd.to_datetime(datetime.today())

    # Split into two categories
    expiring_soon = upcoming[
        (upcoming['Contract_End_Date'] - today).dt.days.between(0, 30)
    ]

    already_expired = upcoming[
        (upcoming['Contract_End_Date'] < today)
    ]

    # ⚠️ Expiring Soon
    if not expiring_soon.empty:
        st.warning("⚠️ The following staff have contracts ending in the next 30 days:")
        df = expiring_soon[['Full_Name', 'CNIC_No', 'Contract_End_Date']].copy()
        df['Contract_End_Date'] = df['Contract_End_Date'].dt.strftime('%Y-%m-%d')
        st.dataframe(df, use_container_width=True)

    else:
        st.success("✅ No contracts expiring in the next 30 days.")

    # ❌ Already Expired
    if not already_expired.empty:
        st.error("❌ The following staff have contracts that have already expired:")
        st.dataframe(already_expired[['Full_Name', 'CNIC_No', 'Contract_End_Date']], use_container_width=True)
    st.markdown("---")

    # ───── Pie Charts ─────
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📍 Province Distribution")
        counts = active_df['Province'].value_counts()
        labels = [f"{k} ({v})" for k, v in counts.items()]
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.pie(counts, labels=labels, autopct='%1.1f%%', startangle=140, textprops={'fontsize': 10})
        ax.axis('equal')
        st.pyplot(fig)
        plt.close()

    with col2:
        st.subheader("⚧️ Gender Distribution")
        counts = active_df['Gender'].value_counts()
        labels = [f"{k} ({v})" for k, v in counts.items()]
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.pie(counts, labels=labels, autopct='%1.1f%%', startangle=140, textprops={'fontsize': 10})
        ax.axis('equal')
        st.pyplot(fig)
        plt.close()

    st.markdown("---")

    # ───── Project Participation ─────
    st.subheader("📌 Project Participation")
    project_counts = {proj: active_df[proj].sum() for proj in PROJECTS}
    fig, ax = plt.subplots(figsize=(10, 4))
    sns.barplot(x=list(project_counts.keys()), y=list(project_counts.values()), palette="Set2", ax=ax)
    ax.set_ylabel("Staff Count")
    ax.set_xlabel("Project")
    ax.set_title("Active Staff by Project")
    ax.grid(axis='y', linestyle='--', alpha=0.6)
    for container in ax.containers:
        ax.bar_label(container, label_type='edge', padding=3, fontsize=9)
    plt.xticks(rotation=30, ha='right')
    st.pyplot(fig)
    plt.close()

    st.markdown("---")     

elif menu == "👥 View Profiles":
    if "view_cnic" in st.session_state and st.session_state.view_cnic:
        st.header("👤 Staff Profile")

        cnic = st.session_state.view_cnic
        profile = active_df[active_df["CNIC_No"].astype(str) == str(cnic)].iloc[0]

        st.subheader(f"👤 Profile: {profile['Full_Name']}")
        st.markdown("---")
        photo_col, info_col = st.columns([1, 3])

        with photo_col:
            if isinstance(profile["Profile_Image"], str) and os.path.exists(profile["Profile_Image"]):
                st.image(profile["Profile_Image"], width=180)
            else:
                st.text("No Image")

        with info_col:
            all_fields = profile.to_dict()
            grid = st.columns(2)
            for i, (key, val) in enumerate(all_fields.items()):
                if isinstance(val, pd.Timestamp):
                    val = val.strftime('%Y-%m-%d')
                val = val if pd.notna(val) else "N/A"
                with grid[i % 2]:
                    st.markdown(f"**{key}**: {val}")

        st.markdown("---")
        if st.button("🔙 Back to List"):
            st.session_state.view_cnic = None
            st.rerun()

    else:
        # ───── Filters and list table ─────
        st.header("👥 View Staff Profiles")

        col1, col2, col3, col4 = st.columns(4)
        province_filter = col1.selectbox("Province", ["All"] + sorted(active_df['Province'].dropna().astype(str).unique()))
        designation_filter = col2.selectbox("Designation", ["All"] + sorted(active_df['Designation'].dropna().astype(str).unique()))
        duty_filter = col3.selectbox("Duty Station", ["All"] + sorted(active_df['District - Duty Station'].dropna().astype(str).unique()))
        reset = col4.button("🔄 Reset Filters")

        empcode_search = st.selectbox("🔍 Search by Employee Code", ["None"] + active_df['Emp_Code'].astype(str).tolist())
        cnic_search = st.selectbox("🔍 Search by CNIC", ["None"] + active_df['CNIC_No'].astype(str).tolist())

        if reset:
            st.session_state.view_cnic = None
            st.rerun()

        filtered_df = active_df.copy()
        if province_filter != "All":
            filtered_df = filtered_df[filtered_df['Province'].astype(str) == province_filter]
        if designation_filter != "All":
            filtered_df = filtered_df[filtered_df['Designation'].astype(str) == designation_filter]
        if duty_filter != "All":
            filtered_df = filtered_df[filtered_df['District - Duty Station'].astype(str) == duty_filter]

        if empcode_search != "None":
            filtered_df = filtered_df[filtered_df['Emp_Code'].astype(str) == empcode_search]
            st.session_state.view_cnic = filtered_df.iloc[0]['CNIC_No']
            st.rerun()
        elif cnic_search != "None":
            filtered_df = filtered_df[filtered_df['CNIC_No'].astype(str) == cnic_search]
            st.session_state.view_cnic = filtered_df.iloc[0]['CNIC_No']
            st.rerun()

        st.markdown(f"### Showing {len(filtered_df)} staff members")

        display_df = filtered_df[[
            "Emp_Code", "Full_Name", "Designation", "Province", "District - Duty Station",
            "Mobile Number", "Email Adresss", "CNIC_No"
        ]].rename(columns={
            "Full_Name": "Name",
            "District - Duty Station": "Duty Station",
            "Mobile Number": "Mobile",
            "Email Adresss": "Email",
            "CNIC_No": "CNIC"
        }).reset_index(drop=True)

        for i, row in display_df.iterrows():
            cols = st.columns([6, 1])
            with cols[0]:
                st.markdown(
                    f"**{row['Name']}** | {row['Designation']} | {row['Duty Station']} | {row['Mobile']} | {row['Email']}"
                )
            with cols[1]:
                if st.button("👁 View", key=f"view_btn_{i}"):
                    st.session_state.view_cnic = row["CNIC"]
                    st.rerun()

# ---------- EDIT EMPLOYEE ----------
elif menu == "✏️ Edit Employee":
    st.header("✏️ Edit Employee")

    # ───── Static Filters ─────
    f1, f2, f3 = st.columns(3)
    name_filter = f1.text_input("Search by Name")
    emp_filter = f2.text_input("Search by Employee Code")
    cnic_filter = f3.text_input("Search by CNIC")

    f4, f5, f6 = st.columns(3)
    province_filter = f4.selectbox("Province", ["All"] + sorted(active_df['Province'].dropna().astype(str).unique()))
    project_filter = f5.selectbox("Project", ["All"] + sorted(active_df['Project'].dropna().astype(str).unique()))
    designation_filter = f6.selectbox("Designation", ["All"] + sorted(active_df['Designation'].dropna().astype(str).unique()))

    filtered_df = active_df.copy()
    if name_filter:
        filtered_df = filtered_df[filtered_df['Full_Name'].str.contains(name_filter, case=False, na=False)]
    if emp_filter:
        filtered_df = filtered_df[filtered_df['Emp_Code'].astype(str).str.contains(emp_filter, na=False)]
    if cnic_filter:
        filtered_df = filtered_df[filtered_df['CNIC_No'].astype(str).str.contains(cnic_filter, na=False)]
    if province_filter != "All":
        filtered_df = filtered_df[filtered_df['Province'].astype(str) == province_filter]
    if project_filter != "All":
        filtered_df = filtered_df[filtered_df['Project'].astype(str) == project_filter]
    if designation_filter != "All":
        filtered_df = filtered_df[filtered_df['Designation'].astype(str) == designation_filter]

    edit_cnic = st.selectbox("Choose CNIC to Edit", filtered_df['CNIC_No'].astype(str).unique() if not filtered_df.empty else [])

    if edit_cnic:
        emp = active_df[active_df['CNIC_No'].astype(str) == edit_cnic].iloc[0]
        with st.form("edit_form"):
            cols = st.columns(2)
            with cols[0]:
                full_name = st.text_input("Full Name", emp.get('Full_Name', ''))
                emp_code = st.text_input("Employee Code", emp.get('Emp_Code', ''))
                cnic_no = st.text_input("CNIC No", emp.get('CNIC_No', ''))
                mobile = st.text_input("Mobile Number", emp.get('Mobile Number', ''))
                email = st.text_input("Email Address", emp.get('Email Adresss', ''))
                dob = st.date_input("Date of Birth", value=pd.to_datetime(emp.get('DOB', pd.Timestamp.now())))
                province = st.text_input("Province", emp.get('Province', ''))
                duty_station = st.text_input("Duty Station", emp.get('District - Duty Station', ''))
                unit = st.text_input("Unit", emp.get('Unit', ''))
                gender = st.selectbox("Gender", ["Male", "Female", "Other"], index=["Male", "Female", "Other"].index(emp.get('Gender', 'Male')) if emp.get('Gender') in ["Male", "Female", "Other"] else 0)
                marital_status = st.selectbox("Marital Status", ["Single", "Married", "Divorced", "Widowed"], index=["Single", "Married", "Divorced", "Widowed"].index(emp.get('Marital Status', 'Single')) if emp.get('Marital Status') in ["Single", "Married", "Divorced", "Widowed"] else 0)
                education = st.text_input("Education", emp.get('Education', ''))

            with cols[1]:
                designation = st.text_input("Designation", emp.get('Designation', ''))
                project = st.text_input("Project", emp.get('Project', ''))
                contract_start = st.date_input("Contract Start Date", value=pd.to_datetime(emp.get('Contract_Start_Date', pd.Timestamp.now())))
                contract_end = st.date_input("Contract End Date", value=pd.to_datetime(emp.get('Contract_End_Date', pd.Timestamp.now())))
                father_name = st.text_input("Father Name", emp.get('Father Name', ''))
                postal_address = st.text_area("Postal Address", emp.get('Postal Address', ''))
                bank_name = st.text_input("Bank Name", emp.get('Bank Name', ''))
                branch_name = st.text_input("Branch Name", emp.get('Branch Name', ''))
                branch_code = st.text_input("Branch Code", emp.get('Branch Code', ''))
                account_number = st.text_input("Bank Account Number", emp.get('Bank Account Number', ''))
                profile_img = st.file_uploader("Upload Profile Image", type=["png", "jpg", "jpeg"])

            st.markdown("### ✅ Active Projects")
            active_projects = {}
            for proj in PROJECTS:
                active_projects[proj] = st.checkbox(proj, value=bool(emp.get(proj, False)))

            remarks = st.text_area("Remarks", emp.get("Remarks", ""))

            submitted = st.form_submit_button("Update Record")
            if submitted:
                idx = active_df[active_df['CNIC_No'].astype(str) == edit_cnic].index[0]
                active_df.at[idx, 'Full_Name'] = full_name
                active_df.at[idx, 'Mobile Number'] = mobile
                active_df.at[idx, 'Email Adresss'] = email
                active_df.at[idx, 'District - Duty Station'] = duty_station
                active_df.at[idx, 'Designation'] = designation
                active_df.at[idx, 'Unit'] = unit
                active_df.at[idx, 'Project'] = project
                active_df.at[idx, 'Province'] = province
                active_df.at[idx, 'DOB'] = dob
                active_df.at[idx, 'Contract_Start_Date'] = contract_start
                active_df.at[idx, 'Contract_End_Date'] = contract_end
                active_df.at[idx, 'Emp_Code'] = emp_code
                active_df.at[idx, 'CNIC_No'] = cnic_no
                active_df.at[idx, 'Father Name'] = father_name
                active_df.at[idx, 'Postal Address'] = postal_address
                active_df.at[idx, 'Gender'] = gender
                active_df.at[idx, 'Marital Status'] = marital_status
                active_df.at[idx, 'Education'] = education
                active_df.at[idx, 'Bank Name'] = bank_name
                active_df.at[idx, 'Branch Name'] = branch_name
                active_df.at[idx, 'Branch Code'] = branch_code
                active_df.at[idx, 'Bank Account Number'] = account_number
                active_df.at[idx, 'Remarks'] = remarks

                for proj in PROJECTS:
                    active_df.at[idx, proj] = active_projects[proj]

                if profile_img:
                    img_path = os.path.join(PROFILE_IMG_DIR, f"{edit_cnic}.png")
                    with open(img_path, "wb") as f:
                        f.write(profile_img.read())
                    active_df.at[idx, "Profile_Image"] = img_path
                else:
                    current_path = active_df.at[idx, "Profile_Image"]
                    if not isinstance(current_path, str):
                        active_df.at[idx, "Profile_Image"] = ""

                save_data(active_df, resigned_df)
                st.success("✅ Record updated successfully")
                st.rerun()


# ---------- CLOSE CONTRACT ----------
elif menu == "📤 Close Contract":
    st.header("📤 Close Contract")

    # =======================
    #     🔍 FILTERS
    # =======================
    st.subheader("Filters")
    f1, f2, f3 = st.columns(3)
    name_filter = f1.text_input("Search by Name")
    emp_filter = f2.text_input("Search by Employee Code (PERN)")
    dummy = f3.markdown("<br>", unsafe_allow_html=True)  # maintain layout

    f4, f5, f6 = st.columns(3)
    province_filter = f4.selectbox("Province", ["All"] + sorted(active_df['Province'].dropna().astype(str).unique()))
    district_filter = f5.selectbox("District (Duty Station)", ["All"] + sorted(active_df['District - Duty Station'].dropna().astype(str).unique()))
    designation_filter = f6.selectbox("Designation", ["All"] + sorted(active_df['Designation'].dropna().astype(str).unique()))

    # Apply filters
    filtered_active = active_df.copy()

    if name_filter:
        filtered_active = filtered_active[filtered_active['Full_Name'].str.contains(name_filter, case=False, na=False)]
    if emp_filter:
        filtered_active = filtered_active[filtered_active['Emp_Code'].astype(str).str.contains(emp_filter, na=False)]
    if province_filter != "All":
        filtered_active = filtered_active[filtered_active['Province'].astype(str) == province_filter]
    if district_filter != "All":
        filtered_active = filtered_active[filtered_active['District - Duty Station'].astype(str) == district_filter]
    if designation_filter != "All":
        filtered_active = filtered_active[filtered_active['Designation'].astype(str) == designation_filter]

    st.write(f"🔎 **Filtered Staff Count:** {len(filtered_active)}")

    # =======================
    #     SINGLE CLOSE
    # =======================
    st.subheader("Single Staff Close")

    if filtered_active.empty:
        st.warning("No staff match your filters.")
    else:
        single_cnic = st.selectbox("Select CNIC to Close Contract", filtered_active['CNIC_No'].astype(str).unique(), key="close_single")
        remarks_input = st.text_area("📝 Reason for Closing Contract (Remarks)", placeholder="Example: Contract ended, poor performance, disciplinary issue, resignation, etc.")

        if st.button("Close Contract", key="close_button"):
            closing_record = active_df[active_df['CNIC_No'].astype(str) == single_cnic].copy()

            if not closing_record.empty:
                active_df = active_df[active_df['CNIC_No'].astype(str) != single_cnic]

                # Add Remarks
                closing_record["Remarks"] = remarks_input if remarks_input.strip() != "" else "No remarks provided"

                resigned_df = pd.concat([resigned_df, closing_record], ignore_index=True)
                save_data(active_df, resigned_df)
                st.success("Contract closed and staff moved to Inactive list.")
            else:
                st.warning("CNIC not found in active records.")

    st.markdown("---")

    # =======================
    #     BULK CLOSE
    # =======================
    st.subheader("📥 Bulk Close Contracts")

    sample_bulk = pd.DataFrame({
        "CNIC_No": ["1234567890123", "9876543210123"],
        "Remarks": ["Contract Ended", "Performance Issue"]
    })

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        sample_bulk.to_excel(writer, index=False, sheet_name="To_Close")

    st.download_button(
        "📄 Download Bulk Close Template",
        data=buffer.getvalue(),
        file_name="bulk_close_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    uploaded_file = st.file_uploader("Upload Filled Template for Bulk Contract Closure", type=["xlsx"], key="bulk_close_upload")

    if uploaded_file:
        try:
            df_bulk = pd.read_excel(uploaded_file)

            if "Remarks" not in df_bulk.columns:
                df_bulk["Remarks"] = "No remarks provided"

            cnic_list = df_bulk['CNIC_No'].astype(str).tolist()

            to_close = active_df[active_df['CNIC_No'].astype(str).isin(cnic_list)].copy()

            if not to_close.empty:
                # Merge Remarks from uploaded file
                to_close = to_close.merge(df_bulk[['CNIC_No', 'Remarks']], on="CNIC_No", how="left")

                active_df = active_df[~active_df['CNIC_No'].astype(str).isin(cnic_list)]
                resigned_df = pd.concat([resigned_df, to_close], ignore_index=True)
                save_data(active_df, resigned_df)
                st.success(f"Successfully closed contracts for {len(to_close)} staff.")
            else:
                st.warning("No matching CNICs found in active staff.")
        except Exception as e:
            st.error(f"Error processing file: {e}")


# ---------- ADD / IMPORT STAFF ----------
elif menu == "➕ Add Staff":
    st.header("➕ Add or Import Staff")

    # ------------------------------------------------------
    #                🔹 SINGLE ENTRY FORM
    # ------------------------------------------------------
    st.subheader("Single Entry")
    with st.form("add_staff_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            emp_code = st.text_input("Employee Code")
            full_name = st.text_input("Full Name")
            cnic = st.text_input("CNIC No")
            mobile = st.text_input("Mobile Number")
            email = st.text_input("Email Address")
            dob = st.date_input("Date of Birth")
        with c2:
            designation = st.text_input("Designation")
            unit = st.text_input("Unit")
            project = st.text_input("Project")
            province = st.text_input("Province")
            duty_station = st.text_input("Duty Station")
            contract_start = st.date_input("Contract Start Date")
            contract_end = st.date_input("Contract End Date")

        st.markdown("### ✅ Active Projects")
        active_projects = {proj: st.checkbox(proj) for proj in PROJECTS}

        submit_btn = st.form_submit_button("Add Staff")

        if submit_btn:
            cnic_str = str(cnic).strip()

            # 🔍 1. Check CNIC in inactive staff
            if cnic_str in resigned_df['CNIC_No'].astype(str).values:
                record = resigned_df[resigned_df['CNIC_No'].astype(str) == cnic_str].iloc[0]
                remarks_text = record.get("Remarks", "No remarks")

                st.error(
                    f"""
                    🚫 **This CNIC belongs to an inactive employee and cannot be added again.**

                    **Name:** {record.get('Full_Name', 'Unknown')}  
                    **Reason (Remarks):** `{remarks_text}`
                    """
                )
                st.stop()

            # 🔍 2. Check CNIC in active staff
            if cnic_str in active_df['CNIC_No'].astype(str).values:
                st.error("⚠️ This CNIC already exists in the Active Staff list. Cannot add duplicate.")
                st.stop()

            # 3️⃣ If safe → allow adding
            new_row = {
                "Emp_Code": emp_code,
                "Full_Name": full_name,
                "CNIC_No": cnic_str,
                "Mobile Number": mobile,
                "Email Adresss": email,
                "DOB": dob,
                "Designation": designation,
                "Unit": unit,
                "Project": project,
                "Province": province,
                "District - Duty Station": duty_station,
                "Contract_Start_Date": contract_start,
                "Contract_End_Date": contract_end,
                "Profile_Image": ""
            }
            for proj in PROJECTS:
                new_row[proj] = active_projects[proj]

            active_df = pd.concat([active_df, pd.DataFrame([new_row])], ignore_index=True)
            save_data(active_df, resigned_df)
            st.success("Staff added successfully.")

    # ------------------------------------------------------
    #                🔹 BULK IMPORT
    # ------------------------------------------------------
    st.markdown("---")
    st.subheader("📥 Bulk Import")

    # Sample template
    sample_bulk = pd.DataFrame({
        "Emp_Code": ["EMP001"],
        "Full_Name": ["Ali Khan"],
        "CNIC_No": ["1234567890123"],
        "Mobile Number": ["03001234567"],
        "Email Adresss": ["ali@example.com"],
        "DOB": [pd.Timestamp("1990-01-01")],
        "Designation": ["Officer"],
        "Unit": ["IT"],
        "Project": ["PMS"],
        "Province": ["Punjab"],
        "District - Duty Station": ["Lahore"],
        "Contract_Start_Date": [pd.Timestamp("2024-01-01")],
        "Contract_End_Date": [pd.Timestamp("2024-12-31")]
    })
    for proj in PROJECTS:
        sample_bulk[proj] = [False]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        sample_bulk.to_excel(writer, index=False, sheet_name="Staff")

    st.download_button(
        "📄 Download Import Template",
        data=buffer.getvalue(),
        file_name="staff_import_template.xlsx"
    )

    uploaded_file = st.file_uploader("Upload Filled Template", type=["xlsx"], key="staff_import")

    if uploaded_file:
        try:
            import_df = pd.read_excel(uploaded_file)

            # Ensure project columns exist
            for proj in PROJECTS:
                if proj not in import_df.columns:
                    import_df[proj] = False

            # Make sure Profile_Image column exists
            if "Profile_Image" not in import_df.columns:
                import_df["Profile_Image"] = ""

            # --------- NEW VALIDATION FOR BULK IMPORT ---------
            inactive_cnic_list = resigned_df['CNIC_No'].astype(str).tolist()
            active_cnic_list = active_df['CNIC_No'].astype(str).tolist()

            skip_rows = []
            valid_rows = []

            for _, row in import_df.iterrows():
                cnic_str = str(row["CNIC_No"]).strip()

                # Check in inactive staff
                if cnic_str in inactive_cnic_list:
                    remarks = resigned_df[resigned_df["CNIC_No"].astype(str) == cnic_str]["Remarks"].iloc[0]
                    skip_rows.append((cnic_str, row.get("Full_Name", "Unknown"), remarks))
                    continue

                # Check in active staff
                if cnic_str in active_cnic_list:
                    skip_rows.append((cnic_str, row.get("Full_Name", "Unknown"), "Already Active"))
                    continue

                valid_rows.append(row)

            if skip_rows:
                st.warning("⚠️ Some rows were skipped:\n")
                for cnic_val, name_val, reason in skip_rows:
                    st.write(f"- **{name_val}** ({cnic_val}) → `{reason}`")

            if valid_rows:
                valid_df = pd.DataFrame(valid_rows)
                active_df = pd.concat([active_df, valid_df], ignore_index=True)
                save_data(active_df, resigned_df)
                st.success(f"Imported {len(valid_rows)} new staff.")
            else:
                st.error("🚫 No valid rows found for import. Nothing was added.")

        except Exception as e:
            st.error(f"Error importing data: {e}")

# ---------- DELETE STAFF ----------
elif menu == "❌ Delete Staff":
    st.header("🗑️ Delete Staff")

    st.subheader("Single Staff Deletion")
    del_cnic = st.selectbox("Select CNIC to Delete", active_df['CNIC_No'].astype(str).unique(), key="del_single")
    if st.button("Delete Selected Staff"):
        active_df = active_df[active_df['CNIC_No'].astype(str) != del_cnic]
        save_data(active_df, resigned_df)
        st.success("Staff deleted successfully.")

    st.markdown("---")
    st.subheader("📥 Bulk Delete by CNICs")

    sample_bulk_delete = pd.DataFrame({
        "CNIC_No": ["1234567890123"]
    })
    delete_buffer = io.BytesIO()
    with pd.ExcelWriter(delete_buffer, engine='xlsxwriter') as writer:
        sample_bulk_delete.to_excel(writer, index=False, sheet_name="Delete")
    st.download_button("📄 Download Bulk Delete Template", data=delete_buffer.getvalue(), file_name="bulk_delete_template.xlsx")

    del_upload = st.file_uploader("Upload Filled Delete Template", type=["xlsx"], key="bulk_delete_upload")
    if del_upload:
        try:
            del_df = pd.read_excel(del_upload)
            del_cnic_list = del_df["CNIC_No"].astype(str).tolist()
            before_count = len(active_df)
            active_df = active_df[~active_df['CNIC_No'].astype(str).isin(del_cnic_list)]
            save_data(active_df, resigned_df)
            removed = before_count - len(active_df)
            st.success(f"Deleted {removed} staff records.")
        except Exception as e:
            st.error(f"Error in bulk deletion: {e}")
# ---------- DOWNLOAD DATA ----------
elif menu == "📥 Download Data":
    st.header("📥 Download Staff Data")

    st.subheader("Download Active Staff List")
    active_buffer = io.BytesIO()
    with pd.ExcelWriter(active_buffer, engine='xlsxwriter') as writer:
        active_df.to_excel(writer, index=False, sheet_name=ACTIVE_SHEET)
    st.download_button("📄 Download Active Staff Excel", data=active_buffer.getvalue(), file_name="active_staff.xlsx")

    st.subheader("Download Inactive (Resigned) Staff List")
    resigned_buffer = io.BytesIO()
    with pd.ExcelWriter(resigned_buffer, engine='xlsxwriter') as writer:
        resigned_df.to_excel(writer, index=False, sheet_name=RESIGNED_SHEET)
    st.download_button("📄 Download Inactive Staff Excel", data=resigned_buffer.getvalue(), file_name="inactive_staff.xlsx")
# ---------- INACTIVE STAFF ----------
elif menu == "🚫 Inactive Staff":
    st.header("🚫 Inactive / Resigned Staff")

    if resigned_df.empty:
        st.info("No inactive staff found.")
    else:
        with st.expander("🔎 Filter Inactive Records"):
            col1, col2 = st.columns(2)
            with col1:
                province_filter = st.selectbox("Filter by Province", ["All"] + sorted(resigned_df['Province'].dropna().astype(str).unique()), key="res_province")
            with col2:
                designation_filter = st.selectbox("Filter by Designation", ["All"] + sorted(resigned_df['Designation'].dropna().astype(str).unique()), key="res_designation")

        filtered_df = resigned_df.copy()
        if province_filter != "All":
            filtered_df = filtered_df[filtered_df['Province'].astype(str) == province_filter]
        if designation_filter != "All":
            filtered_df = filtered_df[filtered_df['Designation'].astype(str) == designation_filter]

        st.dataframe(filtered_df.sort_values(by="Full_Name"), use_container_width=True, height=500)
        st.write(f"Total Inactive Records: {len(filtered_df)}")

        reactivate_cnic = st.selectbox("Select CNIC to Re-activate", filtered_df['CNIC_No'].astype(str).unique())
        if st.button("♻️ Re-activate Selected Staff"):
            reactivated = resigned_df[resigned_df['CNIC_No'].astype(str) == reactivate_cnic].copy()
            resigned_df = resigned_df[resigned_df['CNIC_No'].astype(str) != reactivate_cnic]
            active_df = pd.concat([active_df, reactivated], ignore_index=True)
            save_data(active_df, resigned_df)
            st.success("Staff successfully reactivated and moved to Active list.")
elif menu == "📆 Attendance":
    st.header("📆 Attendance Tab")

    import openpyxl
    from openpyxl.styles import Alignment

    def generate_attendance_df(start_date, end_date, in_time, out_time):
        dates = pd.date_range(start=start_date, end=end_date)
        data = []
        for i, date in enumerate(dates):
            in_dt = datetime.strptime(in_time, "%H:%M")
            out_dt = datetime.strptime(out_time, "%H:%M")
            work_hours = (out_dt - in_dt).seconds // 3600
            data.append({
                "Sr#": i + 1,
                "Date": date.strftime('%Y-%m-%d'),
                "Start Time": in_time,
                "End Time": out_time,
                "No. of completed work hours": work_hours
            })
        return data

    def safe_write(ws, cell_ref, value):
        cell = ws[cell_ref]
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.value = value

    st.subheader("📄 Single Staff Attendance")
    search_by = st.radio("Search Staff By", ["CNIC", "PERN"], horizontal=True)
    if search_by == "CNIC":
        value = st.selectbox("Select CNIC", active_df['CNIC_No'].astype(str).unique())
        staff_row = active_df[active_df['CNIC_No'].astype(str) == value].iloc[0]
    else:
        value = st.selectbox("Select PERN (Emp Code)", active_df['Emp_Code'].astype(str).unique())
        staff_row = active_df[active_df['Emp_Code'].astype(str) == value].iloc[0]

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date")
        end_date = st.date_input("End Date")
    with col2:
        in_time = st.text_input("Default In-Time (HH:MM)", value="08:00")
        out_time = st.text_input("Default Out-Time (HH:MM)", value="17:00")

    if start_date and end_date and in_time and out_time and start_date <= end_date:
        all_attendance = generate_attendance_df(start_date, end_date, in_time, out_time)
        st.subheader("📋 Select Attendance Dates")

        cols = st.columns(3)
        selected_rows = []
        for i, row in enumerate(all_attendance):
            label = f"{row['Date']} ({row['Start Time']} - {row['End Time']})"
            col = cols[i % 3]
            with col:
                if st.checkbox(label, value=True, key=f"attend_{i}"):
                    selected_rows.append(row)

        if not selected_rows:
            st.warning("No dates selected. Please select at least one.")
        else:
            preview_df = pd.DataFrame(selected_rows)
            st.dataframe(preview_df, use_container_width=True)

            wb = openpyxl.load_workbook("attendance.xlsx")
            ws = wb.active

            safe_write(ws, "B10", staff_row['Full_Name'].split()[-1])
            safe_write(ws, "D10", ' '.join(staff_row['Full_Name'].split()[:-1]))
            safe_write(ws, "F10", str(staff_row['CNIC_No']))
            safe_write(ws, "B11", str(staff_row['Emp_Code']))
            safe_write(ws, "F11", staff_row['Email Adresss'])
            safe_write(ws, "B12", staff_row['Designation'])
            safe_write(ws, "E12", staff_row['District - Duty Station'])
            safe_write(ws, "B13", start_date.strftime('%Y-%m-%d'))
            safe_write(ws, "F13", end_date.strftime('%Y-%m-%d'))

            for i, row in enumerate(selected_rows, start=17):
                ws[f"B{i}"] = row['Date']
                ws[f"C{i}"] = row['Start Time']
                ws[f"D{i}"] = row['End Time']
                ws[f"G{i}"] = row['No. of completed work hours']
                for col in ["B", "C", "D", "G"]:
                    ws[f"{col}{i}"].alignment = Alignment(horizontal="center", vertical="center")

            excel_output = io.BytesIO()
            wb.save(excel_output)
            excel_output.seek(0)

            filename = f"Attendance_{staff_row['Full_Name'].replace(' ', '_')}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}"
            st.download_button(
                "📄 Download Excel",
                data=excel_output.getvalue(),
                file_name=f"{filename}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
